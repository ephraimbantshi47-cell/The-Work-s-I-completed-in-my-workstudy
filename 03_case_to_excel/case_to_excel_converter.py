"""
CASE Statement to Excel Lookup Table Converter
-------------------------------------------------
Parses SQL CASE...WHEN...THEN logic and converts it into a clean, styled
Excel lookup table. Built to turn dense SQL mapping logic (e.g., code ->
category lookups) into a format non-technical staff could read, verify,
and hand off to other departments without needing to open a report
definition.

Sample data below is fictional (generic region/category codes) — in
production this ran against real school/district code mappings pulled
from Cognos report definitions.

Author: Ephraim
"""

import re
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
#  Paste any SQL CASE statement(s) below. Supports:
#    - WHEN [FIELD] IN ('a', 'b') THEN 'label'
#    - WHEN [FIELD] = 'a' THEN 'label'
#    - Multiple CASE...END blocks
#    - ELSE values
# ============================================================

case_input = """
CASE
    WHEN [REGION_CODE] IN ('R001') THEN ('North District')
    WHEN [REGION_CODE] IN ('R002') THEN ('South District')
    WHEN [REGION_CODE] IN ('R003', 'R004') THEN ('East Consortium')
    WHEN [REGION_CODE] IN ('R005', 'R006', 'R007') THEN ('West Regional Schools')
    WHEN [REGION_CODE] IN ('R008') THEN ('Online Academy')
    WHEN [REGION_CODE] = 'HOME' THEN ('Home School')
    WHEN [REGION_CODE] = 'GED' THEN ('GED Program')
    ELSE ('Unclassified / Needs Review')
END
"""

output_file = "case_lookup_output.xlsx"


def parse_case_statements(text):
    rows = []
    segments = re.split(r'\bCASE\b', text, flags=re.IGNORECASE)

    for segment in segments:
        segment = re.sub(r'\bEND\b', '', segment, flags=re.IGNORECASE)

        else_match = re.search(
            r'\bELSE\s+(?:\'([^\']+)\'|\(\'([^\']+)\'\)|\[([^\]]+)\])',
            segment, re.IGNORECASE
        )
        else_value = ''
        if else_match:
            else_value = else_match.group(1) or else_match.group(2) or f"[{else_match.group(3)}]"
            segment = re.sub(r'\bELSE\b.*', '', segment, flags=re.IGNORECASE | re.DOTALL)

        field_match = re.search(r'\[([^\]]+)\]', segment, re.IGNORECASE)
        field_name = field_match.group(1) if field_match else 'CODE'

        when_lines = re.split(r'\bWHEN\b', segment, flags=re.IGNORECASE)

        for line in when_lines:
            line = line.strip()
            if not line:
                continue

            in_match = re.search(
                r'\[[^\]]+\]\s+in\s*\(([^)]+)\)\s+THEN\s+\(?\s*\'([^\']+)\'\s*\)?',
                line, re.IGNORECASE
            )
            if in_match:
                codes_str, label = in_match.group(1), in_match.group(2)
                codes = re.findall(r"'([^']+)'", codes_str)
                for code in codes:
                    rows.append({'Field': field_name, 'Code': code.strip(), 'Label': label.strip(), 'ELSE_Value': else_value})
                continue

            eq_match = re.search(
                r'\[[^\]]+\]\s*=\s*\'([^\']+)\'\s+THEN\s+\(?\s*\'([^\']+)\'\s*\)?',
                line, re.IGNORECASE
            )
            if eq_match:
                code, label = eq_match.group(1), eq_match.group(2)
                rows.append({'Field': field_name, 'Code': code.strip(), 'Label': label.strip(), 'ELSE_Value': else_value})

    return rows


def save_excel(rows, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lookup Table"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ['Code', 'Category']
    data_keys = ['Code', 'Label']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    fill_light = PatternFill("solid", fgColor="EBF3FB")
    fill_white = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, row in enumerate(rows, 2):
        fill = fill_light if row_idx % 2 == 0 else fill_white
        for col, key in enumerate(data_keys, 1):
            cell = ws.cell(row=row_idx, column=col, value=row[key])
            cell.fill = fill
            cell.alignment = Alignment(vertical='center')

    for col_cells in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    wb.save(filepath)


def main():
    print("CASE Statement to Excel Converter")
    print("=" * 40)

    rows = parse_case_statements(case_input)
    if not rows:
        print("No WHEN/THEN rows found. Check formatting of case_input.")
        return

    current_field = None
    for r in rows:
        if r['Field'] != current_field:
            print(f"\n[{r['Field']}]")
            print(f"{'CODE':<15} CATEGORY")
            print("-" * 40)
            current_field = r['Field']
        print(f"{r['Code']:<15} {r['Label']}")

    print(f"\nTotal rows: {len(rows)}")

    script_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
    full_path = os.path.join(script_dir, output_file)
    save_excel(rows, full_path)
    print(f"Excel saved to: {full_path}")


if __name__ == "__main__":
    main()
